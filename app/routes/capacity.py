"""Capacity Report routes – /capacity/"""
import csv
import io
import logging
from flask import Blueprint, render_template, jsonify, request, current_app, Response

bp = Blueprint('capacity', __name__, url_prefix='/capacity')
logger = logging.getLogger(__name__)


@bp.route('/')
def index():
    """Capacity report main page."""
    from app.models import AppSettings
    settings = AppSettings.query.first()
    pure1_configured = bool(settings and settings.pure1_app_id and settings.pure1_private_key)
    return render_template('capacity.html', pure1_configured=pure1_configured)


@bp.route('/api/data')
def api_data():
    """Return aggregated capacity data for all views as JSON."""
    from app.models import StorageSystem
    from app.capacity_service import (
        get_latest_snapshots,
        build_by_storage_art,
        build_by_environment,
        build_by_department,
        build_details,
    )

    systems = StorageSystem.query.filter_by(enabled=True).all()
    snapshots = get_latest_snapshots()

    # Determine staleness: is any snapshot older than 1 hour?
    from datetime import datetime, timedelta
    now = datetime.utcnow()
    stale = not snapshots or any(
        (now - s.fetched_at).total_seconds() > 3600
        for s in snapshots.values()
    )

    # Compute the most-recent fetched_at by comparing datetime objects, then ISO-format it
    last_dt = max((s.fetched_at for s in snapshots.values()), default=None) if snapshots else None

    return jsonify({
        'stale': stale,
        'last_updated': last_dt.isoformat() if last_dt else None,
        'by_storage_art': build_by_storage_art(systems, snapshots),
        'by_environment': build_by_environment(systems, snapshots),
        'by_department': build_by_department(systems, snapshots),
        'details': build_details(systems, snapshots),
    })


@bp.route('/api/history')
def api_history():
    """Return historical capacity data for chart rendering."""
    from app.capacity_service import get_history_data, compute_forecast, get_sod_history_data

    range_param = request.args.get('range', 'all')
    days_map = {'3m': 90, '6m': 180, '1y': 365, '2y': 730}
    days = days_map.get(range_param)

    history = get_history_data(days=days)

    # Attach forecast to each storage art
    for art, art_data in history.items():
        fc = compute_forecast(art_data['labels'], art_data['used'], forecast_days=90)
        art_data['forecast_labels'] = fc['labels']
        art_data['forecast_values'] = fc['values']

    # Attach SoD commercial data to the Block art (if available)
    sod = get_sod_history_data(days=days)
    if sod:
        block_art = 'Block'
        if block_art not in history:
            # Ensure Block entry exists even when no physical history is available,
            # so that SoD-only data can still be displayed in the chart.
            history[block_art] = {
                'labels': [], 'used': [], 'total': [],
                'forecast_labels': [], 'forecast_values': [],
            }
        fc_demand = compute_forecast(sod['labels'], sod['effective_used'], forecast_days=90)
        history[block_art]['sod'] = {
            'labels': sod['labels'],
            'reserved': sod['reserved'],
            'effective_used': sod['effective_used'],
            'on_demand': sod['on_demand'],
            'effective_forecast_labels': fc_demand['labels'],
            'effective_forecast_values': fc_demand['values'],
        }

    return jsonify(history)


@bp.route('/api/subscription-licenses')
def api_subscription_licenses():
    """Return Pure1 subscription-license data from the local cache (Storage on Demand)."""
    from app.models import AppSettings
    from app.sod_service import get_cached_data

    settings = AppSettings.query.first()
    configured = bool(settings and settings.pure1_app_id and settings.pure1_private_key)
    if not configured:
        return jsonify({
            'configured': False,
            'items': [],
            'fetched_at': None,
            'error': 'Pure1 API-Zugangsdaten nicht konfiguriert.',
        })

    data = get_cached_data()
    data['configured'] = True
    return jsonify(data)


@bp.route('/api/sod-refresh', methods=['POST'])
def api_sod_refresh():
    """Trigger an immediate non-blocking SoD data refresh."""
    from app.sod_service import trigger_refresh
    app = current_app._get_current_object()
    trigger_refresh(app)
    return jsonify({'status': 'refresh_triggered'})


@bp.route('/api/refresh', methods=['POST'])
def api_refresh():
    """Trigger an immediate (non-blocking) capacity data refresh."""
    from app.capacity_service import trigger_refresh
    app = current_app._get_current_object()
    trigger_refresh(app)
    return jsonify({'status': 'refresh_triggered'})


@bp.route('/api/export/csv')
def api_export_csv():
    """Export weekly capacity history as CSV file."""
    from app.capacity_service import get_weekly_history_data

    range_param = request.args.get('range', 'all')
    days_map = {'3m': 90, '6m': 180, '1y': 365, '2y': 730}
    days = days_map.get(range_param)

    rows = get_weekly_history_data(days=days)

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=[
        'week', 'week_start', 'storage_art',
        'total_tb', 'used_tb', 'free_tb', 'percent_used',
    ])
    writer.writeheader()
    writer.writerows(rows)

    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename="kapazitaet_wochenweise.csv"'},
    )


@bp.route('/api/export/excel')
def api_export_excel():
    """Export weekly capacity history as Excel (.xlsx) file."""
    from app.capacity_service import get_weekly_history_data
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    range_param = request.args.get('range', 'all')
    days_map = {'3m': 90, '6m': 180, '1y': 365, '2y': 730}
    days = days_map.get(range_param)

    rows = get_weekly_history_data(days=days)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Kapazität wochenweise'

    headers = ['Woche (ISO)', 'Wochenstart', 'Storage Art',
               'Gesamt [TB]', 'Genutzt [TB]', 'Frei [TB]', 'Genutzt [%]']
    header_fill = PatternFill(fill_type='solid', fgColor='0098DB')
    header_font = Font(bold=True, color='FFFFFF')

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row in rows:
        ws.append([
            row['week'],
            row['week_start'],
            row['storage_art'],
            row['total_tb'],
            row['used_tb'],
            row['free_tb'],
            row['percent_used'],
        ])

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="kapazitaet_wochenweise.xlsx"'},
    )


@bp.route('/api/import/history', methods=['POST'])
def api_import_history():
    """Import historical capacity data from a CSV file."""
    from app.models import StorageSystem
    from app.capacity_service import import_history_from_csv

    if 'file' not in request.files:
        return jsonify({'error': 'Keine Datei übermittelt.'}), 400

    f = request.files['file']
    if not f.filename.lower().endswith('.csv'):
        return jsonify({'error': 'Nur CSV-Dateien werden unterstützt.'}), 400

    systems = StorageSystem.query.all()
    system_map = {s.name.lower(): s for s in systems}

    try:
        raw = f.stream.read()
        try:
            text = raw.decode('utf-8-sig')
        except UnicodeDecodeError:
            return jsonify({
                'error': 'CSV-Datei konnte nicht dekodiert werden. Bitte UTF-8-Kodierung verwenden.'
            }), 400
        stream = io.StringIO(text)
        imported, skipped, errors = import_history_from_csv(stream, system_map)
    except Exception as exc:
        logger.exception('History import failed')
        return jsonify({'error': str(exc)}), 500

    return jsonify({
        'imported': imported,
        'skipped': skipped,
        'errors': errors[:20],  # cap error list
    })


@bp.route('/api/import/sod-history-pure1', methods=['POST'])
def api_import_sod_history_pure1():
    """Import historical Storage on Demand data directly from the Pure1 API."""
    from datetime import date as _date
    from app.capacity_service import import_sod_history_from_pure1

    body = request.get_json(silent=True) or {}
    start_str = body.get('start_date', '')
    end_str = body.get('end_date', '')

    if not start_str or not end_str:
        return jsonify({'error': 'start_date und end_date sind Pflichtfelder (YYYY-MM-DD).'}), 400

    try:
        start_date = _date.fromisoformat(start_str)
        end_date = _date.fromisoformat(end_str)
    except ValueError as exc:
        return jsonify({'error': f'Ungültiges Datumsformat: {exc}'}), 400

    if end_date < start_date:
        return jsonify({'error': 'end_date darf nicht vor start_date liegen.'}), 400

    try:
        imported, skipped, errors = import_sod_history_from_pure1(start_date, end_date)
    except RuntimeError as exc:
        return jsonify({'error': str(exc)}), 400
    except Exception as exc:
        logger.exception('SoD Pure1 history import failed')
        return jsonify({'error': str(exc)}), 500

    return jsonify({
        'imported': imported,
        'skipped': skipped,
        'errors': errors[:20],
    })
